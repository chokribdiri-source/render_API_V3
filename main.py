import os
import time
import json
import logging
import threading
from decimal import Decimal, ROUND_DOWN
from typing import Dict, Any
from datetime import datetime
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import FastAPI, Request, HTTPException
from fastapi.responses import FileResponse
from binance.client import Client
from binance.exceptions import BinanceAPIException
from dotenv import load_dotenv

# ==================== CHARGEMENT CONFIGURATION .env ====================
load_dotenv()

# Configuration depuis .env
API_KEY = os.getenv("BINANCE_API_KEY", "")
API_SECRET = os.getenv("BINANCE_API_SECRET", "")
USE_TESTNET = os.getenv("USE_TESTNET", "true").lower() == "true"
PORT = int(os.getenv("PORT", 8000))
POLL_INTERVAL = float(os.getenv("POLL_INTERVAL", 2.0))
LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()

# ⚠️ CHEMINS RELATIFS POUR DÉPLOIEMENT
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATE_FILE_PATH = os.getenv("STATE_FILE_PATH", os.path.join(BASE_DIR, "state.json"))
HISTORY_EXCEL_PATH = os.getenv("HISTORY_EXCEL_PATH", os.path.join(BASE_DIR, "trading_history.xlsx"))

# Vérification des clés API
if not API_KEY or not API_SECRET:
    raise Exception("❌ Clés API manquantes! Configure BINANCE_API_KEY et BINANCE_API_SECRET dans .env")

# Configuration du logging
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

app = FastAPI()

# Initialisation du client Binance
if USE_TESTNET:
    client = Client(API_KEY, API_SECRET, testnet=True)
    logging.info("🔧 Mode TESTNET activé")
else:
    client = Client(API_KEY, API_SECRET)
    logging.info("🚀 Mode LIVE activé - ATTENTION!")

# Ta stratégie de niveaux
LEVELS = [
    {"capital": 1.0,  "leverage": 50, "tp_pct": 0.003, "sl_pct": 0.003},
    {"capital": 2.0,  "leverage": 50, "tp_pct": 0.003, "sl_pct": 0.003},
    {"capital": 4.5,  "leverage": 50, "tp_pct": 0.003, "sl_pct": 0.003},
    {"capital": 9.5,  "leverage": 50, "tp_pct": 0.003, "sl_pct": 0.003},
    {"capital": 16.0, "leverage": 65, "tp_pct": 0.003, "sl_pct": 0.003},
]

# ==================== GESTION D'ÉTAT AVEC VERROUS ====================
state_lock = threading.Lock()
symbol_locks: Dict[str, threading.Lock] = {}

def get_symbol_lock(symbol: str):
    with state_lock:
        if symbol not in symbol_locks:
            symbol_locks[symbol] = threading.Lock()
        return symbol_locks[symbol]

def load_state():
    """Charge l'état depuis le fichier JSON avec verrou"""
    with state_lock:
        try:
            with open(STATE_FILE_PATH, "r") as f:
                return json.load(f)
        except FileNotFoundError:
            logging.info(f"📄 Fichier state.json non trouvé, création: {STATE_FILE_PATH}")
            return {"positions": {}, "processed_alerts": {}}

def save_state(state):
    """Sauvegarde l'état dans le fichier JSON avec verrou"""
    with state_lock:
        with open(STATE_FILE_PATH, "w") as f:
            json.dump(state, f, indent=2)
        logging.info(f"💾 State sauvegardé: {STATE_FILE_PATH}")

# ==================== GESTION HISTORIQUE EXCEL ====================
def init_excel_history():
    """Initialise le fichier Excel avec les en-têtes"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Trading History"
        
        # En-têtes
        headers = [
            "ID", "Date Heure", "Type", "Symbole", "Direction", "Niveau",
            "Prix Entrée", "Quantité", "Capital", "Effet Levier",
            "Prix TP", "Prix SL", "Prix Fermeture", "Type Fermeture",
            "Profit/Loss (USDT)", "Statut", "Order ID", "TP Order ID", "SL Order ID",
            "Niveau Renforcement Suivant", "Durée Position", "Timestamp"
        ]
        
        ws.append(headers)
        
        # Style des en-têtes
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Ajuster la largeur des colonnes
        column_widths = {
            'A': 8, 'B': 20, 'C': 12, 'D': 10, 'E': 10, 'F': 8,
            'G': 12, 'H': 12, 'I': 10, 'J': 12, 'K': 12, 'L': 12,
            'M': 12, 'N': 12, 'O': 15, 'P': 12, 'Q': 15, 'R': 15,
            'S': 15, 'T': 12, 'U': 15, 'V': 20
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
            
        wb.save(HISTORY_EXCEL_PATH)
        logging.info(f"📊 Fichier Excel historique créé: {HISTORY_EXCEL_PATH}")
        
    except Exception as e:
        logging.error(f"❌ Erreur création fichier Excel: {e}")

def load_history():
    """Charge l'historique depuis le fichier Excel"""
    try:
        if not os.path.exists(HISTORY_EXCEL_PATH):
            init_excel_history()
            return []
            
        df = pd.read_excel(HISTORY_EXCEL_PATH)
        return df.to_dict('records')
    except Exception as e:
        logging.error(f"❌ Erreur chargement historique Excel: {e}")
        return []

def add_to_history(entry_type, data):
    """Ajoute une entrée à l'historique Excel"""
    try:
        if not os.path.exists(HISTORY_EXCEL_PATH):
            init_excel_history()
            
        # Charger l'historique existant
        df_existing = pd.read_excel(HISTORY_EXCEL_PATH)
        
        # Calculer la durée de position si fermeture
        duration = ""
        if entry_type == "POSITION_CLOSED":
            open_timestamp = data.get("open_timestamp")
            if open_timestamp:
                try:
                    open_time = datetime.fromisoformat(open_timestamp.replace('Z', '+00:00'))
                    close_time = datetime.now()
                    duration_seconds = (close_time - open_time).total_seconds()
                    hours = int(duration_seconds // 3600)
                    minutes = int((duration_seconds % 3600) // 60)
                    seconds = int(duration_seconds % 60)
                    duration = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                except Exception as e:
                    logging.warning(f"⚠️ Erreur calcul durée: {e}")
        
        # Préparer la nouvelle ligne
        new_row = {
            "ID": len(df_existing) + 1,
            "Date Heure": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Type": entry_type,
            "Symbole": data.get("symbol", ""),
            "Direction": data.get("direction", ""),
            "Niveau": data.get("level", 1),
            "Prix Entrée": data.get("entry_price", 0),
            "Quantité": data.get("quantity", 0),
            "Capital": data.get("capital", 0),
            "Effet Levier": data.get("leverage", 1),
            "Prix TP": data.get("tp_price", 0),
            "Prix SL": data.get("sl_price", 0),
            "Prix Fermeture": data.get("close_price", 0),
            "Type Fermeture": data.get("close_type", ""),
            "Profit/Loss (USDT)": data.get("profit_loss", 0),
            "Statut": "ACTIVE" if entry_type in ["POSITION_OPENED", "REINFORCEMENT_OPENED"] else "CLOSED",
            "Order ID": data.get("order_id", ""),
            "TP Order ID": data.get("tp_order_id", ""),
            "SL Order ID": data.get("sl_order_id", ""),
            "Niveau Renforcement Suivant": data.get("next_reinforcement_level", 1),
            "Durée Position": duration,
            "Timestamp": datetime.now().isoformat()
        }
        
        # Ajouter la nouvelle ligne
        df_new = pd.DataFrame([new_row])
        df_updated = pd.concat([df_existing, df_new], ignore_index=True)
        
        # Sauvegarder avec mise en forme
        with pd.ExcelWriter(HISTORY_EXCEL_PATH, engine='openpyxl', mode='w') as writer:
            df_updated.to_excel(writer, index=False, sheet_name='Trading History')
            
            # Appliquer la mise en forme
            workbook = writer.book
            worksheet = writer.sheets['Trading History']
            
            # Style des en-têtes
            for col in range(1, len(df_updated.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Style des lignes selon le profit/perte
            for row in range(2, len(df_updated) + 2):
                profit_cell = worksheet.cell(row=row, column=15)  # Colonne Profit/Loss
                try:
                    profit_value = float(profit_cell.value) if profit_cell.value else 0
                    if profit_value > 0:
                        profit_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        profit_cell.font = Font(color="006100")
                    elif profit_value < 0:
                        profit_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        profit_cell.font = Font(color="9C0006")
                except:
                    pass
                
                # Style pour la direction
                direction_cell = worksheet.cell(row=row, column=5)  # Colonne Direction
                if direction_cell.value == "BUY":
                    direction_cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                    direction_cell.font = Font(color="0070C0")
                elif direction_cell.value == "SELL":
                    direction_cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    direction_cell.font = Font(color="FF0000")
        
        logging.info(f"📝 Historique Excel mis à jour: {entry_type} - {data.get('symbol', '')}")
        
    except Exception as e:
        logging.error(f"❌ Erreur ajout historique Excel: {e}")

def calculate_pnl(position, close_type, close_price=None):
    """Calcule le profit/perte d'une position"""
    try:
        entry_price = position.get("entry_price", 0)
        quantity = position.get("quantity", 0)
        
        if close_type == "TP":
            level_config = LEVELS[position.get("current_level", 1)-1]
            if position.get("signal").upper() == "BUY":
                close_price = entry_price * (1 + level_config["tp_pct"])
            else:
                close_price = entry_price * (1 - level_config["tp_pct"])
        elif close_type == "SL":
            level_config = LEVELS[position.get("current_level", 1)-1]
            if position.get("signal").upper() == "BUY":
                close_price = entry_price * (1 - level_config["sl_pct"])
            else:
                close_price = entry_price * (1 + level_config["sl_pct"])
        
        # Si close_price est fourni (fermeture manuelle), l'utiliser
        if close_price is None and close_type == "MANUAL":
            close_price = position.get("close_price", entry_price)
        
        if position.get("signal").upper() == "BUY":
            pnl = (close_price - entry_price) * quantity
        else:
            pnl = (entry_price - close_price) * quantity
            
        return round(pnl, 4)
    except Exception as e:
        logging.error(f"❌ Erreur calcul PnL: {e}")
        return 0

# ==================== CALCULS DE QUANTITÉ ====================
SYMBOL_INFO_CACHE = {}

def fetch_symbol_info(symbol: str):
    if symbol in SYMBOL_INFO_CACHE:
        return SYMBOL_INFO_CACHE[symbol]
    info = client.futures_exchange_info()
    for s in info['symbols']:
        if s['symbol'] == symbol:
            SYMBOL_INFO_CACHE[symbol] = s
            return s
    raise Exception(f"Symbole {symbol} non trouvé")

def get_step_size(symbol: str):
    s = fetch_symbol_info(symbol)
    for f in s['filters']:
        if f['filterType'] == 'LOT_SIZE':
            return float(f['stepSize'])
    return 0.0001

def get_price_precision(symbol: str):
    """Récupère la précision de prix pour un symbole"""
    try:
        symbol_info = fetch_symbol_info(symbol)
        for f in symbol_info['filters']:
            if f['filterType'] == 'PRICE_FILTER':
                tick_size = float(f['tickSize'])
                # Calcul du nombre de décimales
                if tick_size < 1:
                    return len(str(tick_size).split('.')[1].rstrip('0'))
                else:
                    return 0
        return 2  # Valeur par défaut
    except Exception as e:
        logging.warning(f"⚠️ Impossible de récupérer la précision prix: {e}")
        return 2

def get_quantity_precision(symbol):
    """Récupère la précision de quantité pour un symbole"""
    try:
        info = client.futures_exchange_info()
        for s in info['symbols']:
            if s['symbol'] == symbol:
                for f in s['filters']:
                    if f['filterType'] == 'LOT_SIZE':
                        step_size = float(f['stepSize'])
                        # Calcul du nombre de décimales
                        if step_size < 1:
                            return len(str(step_size).split('.')[1].rstrip('0'))
                        return 0
        return 3  # Valeur par défaut
    except Exception as e:
        logging.warning(f"⚠️ Impossible de récupérer la précision: {e}")
        return 3

def round_qty(qty: float, step: float):
    step_dec = Decimal(str(step))
    q = Decimal(str(qty))
    rounded = (q // step_dec) * step_dec
    return float(rounded.quantize(step_dec, rounding=ROUND_DOWN))

def calculate_quantity(capital, leverage, price, symbol):
    """Calcule la quantité avec la bonne précision"""
    notional = capital * leverage
    raw_quantity = notional / price
    
    step = get_step_size(symbol)
    quantity = round_qty(raw_quantity, step)
    
    logging.info(f"📊 Calcul quantité: {capital} × {leverage} = {notional} / {price} = {raw_quantity} → {quantity}")
    return quantity

# ==================== GESTION DES ORDRES ====================
def wait_for_order_execution(symbol, order_id, max_attempts=10):
    """Attend que l'ordre soit exécuté et retourne le prix moyen"""
    for i in range(max_attempts):
        try:
            order_status = client.futures_get_order(symbol=symbol, orderId=order_id)
            status = order_status['status']
            avg_price = float(order_status['avgPrice'])
            executed_qty = float(order_status['executedQty'])
            
            logging.info(f"📊 Statut ordre {i+1}/{max_attempts}: {status}, Prix: {avg_price}, Qty exécutée: {executed_qty}")
            
            if status == 'FILLED' and avg_price > 0:
                logging.info(f"🎉 Ordre exécuté! Prix moyen: {avg_price}")
                return avg_price
            elif status in ['CANCELED', 'EXPIRED', 'REJECTED']:
                raise Exception(f"Ordre {status}")
                
        except Exception as e:
            logging.warning(f"⚠️ Erreur vérification ordre: {e}")
        
        time.sleep(1)
    
    # Fallback: utiliser le prix actuel
    ticker = client.futures_symbol_ticker(symbol=symbol)
    current_price = float(ticker['price'])
    logging.info(f"⏰ Timeout, utilisation prix actuel: {current_price}")
    return current_price

def cancel_order(symbol: str, order_id: int):
    """Annule un ordre"""
    try:
        client.futures_cancel_order(symbol=symbol, orderId=order_id)
        logging.info(f"✅ Ordre annulé: {order_id} sur {symbol}")
    except Exception as e:
        logging.warning(f"❌ Échec annulation ordre {order_id}: {e}")

def get_order_status(symbol: str, order_id: int):
    """Récupère le statut d'un ordre"""
    try:
        order = client.futures_get_order(symbol=symbol, orderId=order_id)
        return order.get("status"), order
    except Exception as e:
        logging.debug(f"❌ Échec récupération statut ordre {order_id}: {e}")
        return None, None

def get_position_amount(symbol: str):
    """Vérification simplifiée de la position"""
    try:
        # Méthode alternative: vérifier via les ordres ouverts
        open_orders = client.futures_get_open_orders(symbol=symbol)
        has_tp_sl = any(order['type'] in ['STOP_MARKET', 'TAKE_PROFIT_MARKET'] for order in open_orders)
        
        if has_tp_sl:
            logging.info(f"🔍 Position {symbol} active (TP/SL trouvés)")
            return 1.0  # Retourne une valeur non nulle
        else:
            logging.info(f"🔍 Position {symbol} - Aucun TP/SL trouvé")
            return 0.0
            
    except Exception as e:
        logging.warning(f"⚠️ Erreur vérification position {symbol}: {e}")
        return 1.0  # En cas d'erreur, suppose que la position est active

# ==================== PLACEMENT DES ORDRES AVEC closePosition ====================
def place_tp_sl_orders_with_retry(symbol, signal, entry_price, level_config, max_retries=3):
    """Place les ordres Take Profit et Stop Loss avec retry en cas d'échec"""
    tp_pct = level_config["tp_pct"]
    sl_pct = level_config["sl_pct"]
    
    if signal.upper() == "BUY":
        tp_price = entry_price * (1 + tp_pct)
        sl_price = entry_price * (1 - sl_pct)
        tp_side = "SELL"
        sl_side = "SELL"
    else:
        tp_price = entry_price * (1 - tp_pct)
        sl_price = entry_price * (1 + sl_pct)
        tp_side = "BUY"
        sl_side = "BUY"
    
    # CORRECTION : Utiliser la bonne précision automatiquement
    price_precision = get_price_precision(symbol)
    tp_price = round(tp_price, price_precision)
    sl_price = round(sl_price, price_precision)
    
    logging.info(f"🎯 TP: {tp_price} (précision: {price_precision}), SL: {sl_price}")
    
    # Ordre Take Profit avec closePosition
    tp_order_id = None
    sl_order_id = None
    
    # Placement TP avec retry
    for attempt in range(max_retries):
        try:
            tp_order = client.futures_create_order(
                symbol=symbol,
                side=tp_side,
                type="TAKE_PROFIT_MARKET",
                stopPrice=tp_price,
                closePosition=True,
                timeInForce="GTC"
            )
            tp_order_id = tp_order.get("orderId")
            logging.info(f"✅ TP placé: {tp_order_id}")
            break
        except Exception as e:
            logging.error(f"❌ Erreur placement TP (tentative {attempt+1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                time.sleep(1)
            else:
                logging.error(f"💥 Échec placement TP après {max_retries} tentatives")
    
    # Placement SL avec retry
    for attempt in range(max_retries):
        try:
            sl_order = client.futures_create_order(
                symbol=symbol,
                side=sl_side,
                type="STOP_MARKET",
                stopPrice=sl_price,
                closePosition=True,
                timeInForce="GTC"
            )
            sl_order_id = sl_order.get("orderId")
            logging.info(f"✅ SL placé: {sl_order_id}")
            break
        except Exception as e:
            logging.error(f"❌ Erreur placement SL (tentative {attempt+1}/{max_retries}): {e}")
            if attempt < max_retries - 1:
                time.sleep(1)
            else:
                logging.error(f"💥 Échec placement SL après {max_retries} tentatives")
    
    return tp_order_id, sl_order_id

def place_binance_order(symbol, signal, quantity, level_config):
    """Place un ordre sur Binance avec TP/SL en utilisant closePosition=True"""
    try:
        leverage = level_config["leverage"]
        
        # 1. Définir le levier
        logging.info(f"🔧 Mise à jour levier: {symbol} → {leverage}")
        client.futures_change_leverage(symbol=symbol, leverage=leverage)
        
        # 2. Déterminer le côté de l'ordre
        side = "BUY" if signal.upper() == "BUY" else "SELL"
        
        # 3. Placer l'ordre MARKET
        logging.info(f"🎯 Placement ordre: {side} {quantity} {symbol}")
        order = client.futures_create_order(
            symbol=symbol,
            side=side,
            type='MARKET',
            quantity=quantity
        )
        
        logging.info(f"✅ Ordre créé: {order['orderId']}")
        
        # 4. Attendre l'exécution et obtenir le prix
        entry_price = wait_for_order_execution(symbol, order['orderId'])
        
        # 5. Placer les ordres TP/SL avec closePosition=True ET retry
        tp_order_id, sl_order_id = place_tp_sl_orders_with_retry(symbol, signal, entry_price, level_config)
        
        return order, entry_price, tp_order_id, sl_order_id
        
    except BinanceAPIException as e:
        logging.error(f"❌ Erreur Binance: {e}")
        raise
    except Exception as e:
        logging.error(f"❌ Erreur inattendue: {e}")
        raise

# ==================== MONITORING AVEC DÉLAI DE GRÂCE ====================
def monitor_loop():
    """Boucle de surveillance des positions et ordres TP/SL"""
    logging.info("🔍 Démarrage du monitoring automatique")
    
    while True:
        try:
            state = load_state()
            positions = state.get("positions", {})
            
            for symbol, position in list(positions.items()):
                if not position.get("is_active", True):
                    continue
                
                # DÉLAI DE GRÂCE : Ne pas vérifier les positions de moins de 30 secondes
                position_timestamp = position.get("timestamp", "")
                if position_timestamp:
                    try:
                        position_time = datetime.fromisoformat(position_timestamp.replace('Z', '+00:00'))
                        time_diff = (datetime.now().replace(tzinfo=None) - position_time.replace(tzinfo=None)).total_seconds()
                        
                        if time_diff < 30:
                            logging.debug(f"⏳ Position {symbol} trop récente ({time_diff:.1f}s) - Attente avant vérification")
                            continue
                    except Exception as e:
                        logging.warning(f"⚠️ Erreur calcul délai position: {e}")
                        continue
                
                # Verrou pour éviter les conflits
                lock = get_symbol_lock(symbol)
                if not lock.acquire(blocking=False):
                    continue
                
                try:
                    current_level = position.get("current_level", 1)
                    tp_order_id = position.get("tp_order_id")
                    sl_order_id = position.get("sl_order_id")
                    signal = position.get("signal")
                    entry_price = position.get("entry_price")
                    
                    # Vérifier d'abord les ordres TP/SL (méthode principale)
                    order_triggered = False
                    
                    if tp_order_id:
                        status, _ = get_order_status(symbol, tp_order_id)
                        if status in ("FILLED", "TRIGGERED"):
                            logging.info(f"🎯 TP exécuté pour {symbol} (niveau {current_level})")
                            # Annuler SL
                            if sl_order_id:
                                cancel_order(symbol, sl_order_id)
                            
                            # Ajouter à l'historique
                            history_data = {
                                "symbol": symbol,
                                "direction": signal,
                                "level": current_level,
                                "entry_price": entry_price,
                                "quantity": position.get("quantity"),
                                "close_type": "TAKE_PROFIT",
                                "profit_loss": calculate_pnl(position, "TP"),
                                "next_reinforcement_level": 1,
                                "open_timestamp": position.get("timestamp")
                            }
                            add_to_history("POSITION_CLOSED", history_data)
                            
                            # Fermer la position dans l'état
                            position["is_active"] = False
                            save_state(state)
                            order_triggered = True
                            continue
                    
                    if sl_order_id and not order_triggered:
                        status, _ = get_order_status(symbol, sl_order_id)
                        if status in ("FILLED", "TRIGGERED"):
                            logging.info(f"🛑 SL exécuté pour {symbol} (niveau {current_level})")
                            # Annuler TP
                            if tp_order_id:
                                cancel_order(symbol, tp_order_id)
                            
                            # Ajouter à l'historique
                            history_data = {
                                "symbol": symbol,
                                "direction": signal,
                                "level": current_level,
                                "entry_price": entry_price,
                                "quantity": position.get("quantity"),
                                "close_type": "STOP_LOSS",
                                "profit_loss": calculate_pnl(position, "SL"),
                                "next_reinforcement_level": current_level + 1 if current_level < len(LEVELS) else 1,
                                "open_timestamp": position.get("timestamp")
                            }
                            add_to_history("POSITION_CLOSED", history_data)
                            
                            # Gérer le renforcement
                            handle_reinforcement(symbol, signal, current_level, state, position)
                            order_triggered = True
                            continue
                    
                    # SEULEMENT SI AUCUN ORDRE TP/SL N'A ÉTÉ DÉCLENCHÉ : vérifier position
                    if not order_triggered:
                        position_amount = get_position_amount(symbol)
                        if position_amount == 0 and position.get("is_active", True):
                            # Vérifier que la position a au moins 60 secondes avant nettoyage
                            if time_diff > 60:
                                logging.info(f"📝 Position {symbol} fermée manuellement après {time_diff:.1f}s - Nettoyage")
                                
                                # Récupérer le prix actuel pour le PnL
                                ticker = client.futures_symbol_ticker(symbol=symbol)
                                current_price = float(ticker['price'])
                                
                                # Ajouter à l'historique
                                history_data = {
                                    "symbol": symbol,
                                    "direction": signal,
                                    "level": current_level,
                                    "entry_price": entry_price,
                                    "quantity": position.get("quantity"),
                                    "close_price": current_price,
                                    "close_type": "MANUAL",
                                    "profit_loss": calculate_pnl(position, "MANUAL", current_price),
                                    "next_reinforcement_level": 1,
                                    "open_timestamp": position.get("timestamp")
                                }
                                add_to_history("POSITION_CLOSED", history_data)
                                
                                position["is_active"] = False
                                if tp_order_id:
                                    cancel_order(symbol, tp_order_id)
                                if sl_order_id:
                                    cancel_order(symbol, sl_order_id)
                                save_state(state)
                            else:
                                logging.debug(f"⏳ Position {symbol} trop récente pour nettoyage ({time_diff:.1f}s)")
                        
                finally:
                    lock.release()
                    
        except Exception as e:
            logging.error(f"❌ Erreur dans monitor_loop: {e}")
        
        time.sleep(5)  # Vérifier toutes les 5 secondes

def handle_reinforcement(symbol, signal, current_level, state, position):
    """Prépare le renforcement pour le prochain signal (quelle que soit la direction)"""
    next_level = current_level + 1
    
    if next_level > len(LEVELS):
        logging.info(f"💥 Niveau maximum atteint pour {symbol} - Séquence terminée")
        position["is_active"] = False
        save_state(state)
        return
    
    # Préparer le renforcement sans direction spécifique
    logging.info(f"⏳ Renforcement préparé: {symbol} prochain signal → niveau {next_level}")
    
    # Marquer la position comme inactive mais garder l'info du niveau suivant
    position.update({
        "is_active": False,
        "pending_reinforcement": True,
        "next_level": next_level
    })
    
    save_state(state)

# Démarrer le monitoring
monitor_thread = threading.Thread(target=monitor_loop, daemon=True)
monitor_thread.start()

# ==================== ENDPOINTS FASTAPI ====================
@app.get("/health")
def health():
    return {"status":"ok", "timestamp": datetime.now().isoformat()}

@app.post("/webhook")
async def webhook(request: Request):
    try:
        data = await request.json()
        logging.info(f"📥 Webhook reçu: {data}")
        
        signal = data.get("signal")
        symbol = data.get("symbol", "ETHUSDC")
        price = float(data.get("price", 0))
        
        if not signal or price == 0:
            raise HTTPException(status_code=400, detail="Signal ou prix manquant")
        
        # Verrou pour ce symbole
        lock = get_symbol_lock(symbol)
        if not lock.acquire(timeout=10):
            raise HTTPException(status_code=429, detail="Symbole occupé")
        
        try:
            state = load_state()
            positions = state.get("positions", {})
            
            # VÉRIFIER SI RENFORCEMENT EN ATTENTE (quelle que soit la direction)
            if symbol in positions:
                position = positions[symbol]
                if position.get("pending_reinforcement", False):
                    next_level = position.get("next_level", 1)
                    
                    # 🔥 OUVRIR DANS LA DIRECTION DU NOUVEAU SIGNAL, AU NIVEAU SUIVANT
                    logging.info(f"🎯 Renforcement activé: {symbol} niveau {next_level} - Direction: {signal}")
                    
                    # Ouvrir la position au niveau suivant avec la NOUVELLE direction
                    level_config = LEVELS[next_level - 1]
                    capital = level_config["capital"]
                    leverage = level_config["leverage"]
                    quantity = calculate_quantity(capital, leverage, price, symbol)
                    
                    if quantity <= 0:
                        raise HTTPException(status_code=400, detail="Quantité invalide")
                    
                    # Placer l'ordre de renforcement avec la NOUVELLE direction
                    order_result, entry_price, tp_order_id, sl_order_id = place_binance_order(
                        symbol, signal, quantity, level_config
                    )
                    
                    # Ajouter à l'historique
                    history_data = {
                        "symbol": symbol,
                        "direction": signal,
                        "level": next_level,
                        "entry_price": entry_price,
                        "quantity": quantity,
                        "capital": capital,
                        "leverage": leverage,
                        "tp_price": entry_price * (1 + level_config["tp_pct"]) if signal.upper() == "BUY" else entry_price * (1 - level_config["tp_pct"]),
                        "sl_price": entry_price * (1 - level_config["sl_pct"]) if signal.upper() == "BUY" else entry_price * (1 + level_config["sl_pct"]),
                        "order_id": order_result['orderId'],
                        "tp_order_id": tp_order_id,
                        "sl_order_id": sl_order_id,
                        "previous_level": next_level - 1,
                        "next_reinforcement_level": next_level + 1 if next_level < len(LEVELS) else 1
                    }
                    add_to_history("REINFORCEMENT_OPENED", history_data)
                    
                    # Mettre à jour l'état
                    position.update({
                        "is_active": True,
                        "pending_reinforcement": False,
                        "current_level": next_level,
                        "signal": signal,  # 🔥 Nouvelle direction
                        "quantity": quantity,
                        "entry_price": entry_price,
                        "capital": capital,
                        "leverage": leverage,
                        "order_id": order_result['orderId'],
                        "tp_order_id": tp_order_id,
                        "sl_order_id": sl_order_id,
                        "timestamp": datetime.now().isoformat()
                    })
                    save_state(state)
                    
                    return {
                        "status": "success", 
                        "message": f"Renforcement {signal} (Niveau {next_level})",
                        "details": {
                            "symbol": symbol,
                            "quantity": quantity,
                            "entry_price": entry_price,
                            "capital": capital,
                            "leverage": leverage,
                            "order_id": order_result['orderId'],
                            "current_level": next_level
                        }
                    }
            
            # VÉRIFICATION DES DOUBLONS (code existant)
            alert_id = f"{symbol}_{signal}_{data.get('time', '')}"
            processed = state.setdefault("processed_alerts", {})
            if alert_id in processed:
                return {"status": "ignored", "reason": "duplicate_alert"}
            processed[alert_id] = int(time.time())
            
            # VÉRIFIER SI POSITION ACTIVE (code existant)
            if symbol in state.get("positions", {}):
                position = state["positions"][symbol]
                if position.get("is_active", True):
                    position_amount = get_position_amount(symbol)
                    if position_amount != 0:
                        return {"status": "ignored", "reason": "position_already_open"}
                    else:
                        # Nettoyer l'état si position fermée
                        del state["positions"][symbol]
            
            # OUVERTURE NOUVELLE POSITION (niveau 1) - code existant
            level_config = LEVELS[0]
            capital = level_config["capital"]
            leverage = level_config["leverage"]
            quantity = calculate_quantity(capital, leverage, price, symbol)
            
            if quantity <= 0:
                raise HTTPException(status_code=400, detail="Quantité invalide")
            
            # Placer l'ordre
            order_result, entry_price, tp_order_id, sl_order_id = place_binance_order(
                symbol, signal, quantity, level_config
            )
            
            # Ajouter à l'historique
            history_data = {
                "symbol": symbol,
                "direction": signal,
                "level": 1,
                "entry_price": entry_price,
                "quantity": quantity,
                "capital": capital,
                "leverage": leverage,
                "tp_price": entry_price * (1 + level_config["tp_pct"]) if signal.upper() == "BUY" else entry_price * (1 - level_config["tp_pct"]),
                "sl_price": entry_price * (1 - level_config["sl_pct"]) if signal.upper() == "BUY" else entry_price * (1 + level_config["sl_pct"]),
                "order_id": order_result['orderId'],
                "tp_order_id": tp_order_id,
                "sl_order_id": sl_order_id,
                "next_reinforcement_level": 2
            }
            add_to_history("POSITION_OPENED", history_data)
            
            # Sauvegarder l'état
            state["positions"][symbol] = {
                "signal": signal,
                "current_level": 1,
                "is_active": True,
                "quantity": quantity,
                "entry_price": entry_price,
                "capital": capital,
                "leverage": leverage,
                "order_id": order_result['orderId'],
                "tp_order_id": tp_order_id,
                "sl_order_id": sl_order_id,
                "alert_id": alert_id,
                "timestamp": datetime.now().isoformat(),
                "pending_reinforcement": False,
                "next_level": 1  # 🔥 Initialiser le niveau suivant
            }
            save_state(state)
            
            return {
                "status": "success", 
                "message": f"Position {signal} ouverte (Niveau 1)",
                "details": {
                    "symbol": symbol,
                    "quantity": quantity,
                    "entry_price": entry_price,
                    "capital": capital,
                    "leverage": leverage,
                    "order_id": order_result['orderId'],
                    "current_level": 1
                }
            }
            
        finally:
            lock.release()
            
    except Exception as e:
        logging.error(f"❌ Erreur webhook: {str(e)}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/")
async def root_post(request: Request):
    """Accepte les POST sur la racine"""
    try:
        logging.info("🔄 Requête reçue sur la racine")
        return await webhook(request)
    except Exception as e:
        logging.error(f"❌ Erreur route racine: {str(e)}")
        return {"status": "error", "message": str(e)}

@app.get("/")
async def root():
    return {"message": "Bot Trading Webhook - Monitoring automatique activé"}

@app.get("/state")
async def get_state():
    """Endpoint pour voir l'état actuel"""
    return load_state()

@app.get("/history")
async def get_history(limit: int = 50):
    """Endpoint pour voir l'historique des trades"""
    history = load_history()
    return {"history": history[-limit:]}

@app.get("/history/excel")
async def download_history():
    """Endpoint pour télécharger le fichier Excel"""
    return FileResponse(
        path=HISTORY_EXCEL_PATH,
        filename="trading_history.xlsx",
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.get("/history/stats")
async def get_history_stats():
    """Statistiques de l'historique"""
    history = load_history()
    
    if not history:
        return {
            "total_trades": 0,
            "total_profit": 0,
            "winning_trades": 0,
            "losing_trades": 0,
            "tp_count": 0,
            "sl_count": 0,
            "manual_count": 0,
            "win_rate": 0
        }
    
    df = pd.DataFrame(history)
    closed_positions = df[df['Statut'] == 'CLOSED']
    
    if closed_positions.empty:
        return {
            "total_trades": 0,
            "total_profit": 0,
            "winning_trades": 0,
            "losing_trades": 0,
            "tp_count": 0,
            "sl_count": 0,
            "manual_count": 0,
            "win_rate": 0
        }
    
    stats = {
        "total_trades": len(closed_positions),
        "total_profit": closed_positions['Profit/Loss (USDT)'].sum(),
        "winning_trades": len(closed_positions[closed_positions['Profit/Loss (USDT)'] > 0]),
        "losing_trades": len(closed_positions[closed_positions['Profit/Loss (USDT)'] < 0]),
        "tp_count": len(closed_positions[closed_positions['Type Fermeture'] == 'TAKE_PROFIT']),
        "sl_count": len(closed_positions[closed_positions['Type Fermeture'] == 'STOP_LOSS']),
        "manual_count": len(closed_positions[closed_positions['Type Fermeture'] == 'MANUAL'])
    }
    
    if stats["total_trades"] > 0:
        stats["win_rate"] = round((stats["winning_trades"] / stats["total_trades"]) * 100, 2)
    else:
        stats["win_rate"] = 0
        
    return stats

@app.delete("/reset")
async def reset_state():
    """Endpoint pour réinitialiser l'état"""
    state = {"positions": {}, "processed_alerts": {}}
    save_state(state)
    return {"status": "reset", "message": "État réinitialisé"}

@app.get("/balance")
async def get_balance():
    """Vérifie le solde du compte"""
    try:
        balance = client.futures_account_balance()
        usdt_balance = next((item for item in balance if item['asset'] == 'USDT'), None)
        return {"balance": usdt_balance}
    except Exception as e:
        return {"error": str(e)}

@app.get("/orders")
async def get_orders(symbol: str = "ETHUSDC"):
    """Vérifie les ordres ouverts"""
    try:
        orders = client.futures_get_open_orders(symbol=symbol)
        return {"symbol": symbol, "open_orders": orders}
    except Exception as e:
        return {"error": str(e)}

@app.get("/check/{symbol}")
async def check_position(symbol: str = "ETHUSDC"):
    """Vérification manuelle par prix (backup)"""
    try:
        ticker = client.futures_symbol_ticker(symbol=symbol)
        current_price = float(ticker['price'])
        
        state = load_state()
        if symbol not in state.get("positions", {}):
            return {"status": "NO_POSITION"}
        
        position = state["positions"][symbol]
        if not position.get("is_active", True):
            return {"status": "POSITION_CLOSED"}
        
        return {
            "symbol": symbol,
            "current_price": current_price,
            "position_active": True,
            "level": position.get("current_level", 1),
            "entry_price": position.get("entry_price"),
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        return {"status": "ERROR", "message": str(e)}

@app.get("/precision/{symbol}")
async def check_precision(symbol: str):
    """Vérifie la précision pour un symbole"""
    try:
        price_precision = get_price_precision(symbol)
        quantity_precision = get_quantity_precision(symbol)
        step_size = get_step_size(symbol)
        
        return {
            "symbol": symbol,
            "price_precision": price_precision,
            "quantity_precision": quantity_precision,
            "step_size": step_size
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/levels")
async def get_levels():
    """Affiche les niveaux de la stratégie"""
    return {
        "strategy": "Renforcement progressif avec monitoring automatique",
        "levels": LEVELS,
        "total_levels": len(LEVELS),
        "total_capital": sum(level["capital"] for level in LEVELS)
    }

if __name__ == "__main__":
    import uvicorn
    logging.info("🚀 Démarrage du bot avec monitoring automatique")
    logging.info(f"📁 State file: {STATE_FILE_PATH}")
    logging.info(f"📊 History Excel file: {HISTORY_EXCEL_PATH}")
    uvicorn.run(app, host="0.0.0.0", port=PORT)